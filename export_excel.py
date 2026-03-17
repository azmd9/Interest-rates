"""
Export CSV curve data to a single Excel workbook with one sheet per curve.

Each daily run appends new rows below existing data. The workbook is
overwritten each time, rebuilt from the full CSV history.
"""

import csv
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, numbers

DATA_DIR = "data"
OUTPUT_FILE = os.path.join(DATA_DIR, "interest_rates.xlsx")

SHEETS = {
    "1M EURIBOR": "1M_EURIBOR.csv",
    "3M EURIBOR": "3M_EURIBOR.csv",
    "3M SONIA": "3M_SONIA.csv",
    "6M EURIBOR": "6M_EURIBOR.csv",
}

HEADER_FONT = Font(bold=True)


def read_csv(filepath):
    """Read a CSV and return header + rows."""
    with open(filepath, newline="") as f:
        reader = csv.reader(f)
        header = next(reader)
        rows = list(reader)
    return header, rows


def populate_sheet(ws, header, rows):
    """Write header and data rows into a worksheet."""
    # Header row
    for col, value in enumerate(header, 1):
        cell = ws.cell(row=1, column=col, value=value)
        cell.font = HEADER_FONT

    # Data rows
    for row_idx, row in enumerate(rows, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            # Try to parse as float for the rate column
            if col_idx == len(header):  # last column is "rate"
                try:
                    cell.value = float(value)
                    cell.number_format = '0.00000%'
                except ValueError:
                    cell.value = value
            else:
                cell.value = value

    # Auto-width columns
    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 2


def main():
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    for sheet_name, csv_file in SHEETS.items():
        csv_path = os.path.join(DATA_DIR, csv_file)
        if not os.path.isfile(csv_path):
            print(f"  Skipping {sheet_name}: {csv_path} not found")
            continue

        header, rows = read_csv(csv_path)
        ws = wb.create_sheet(title=sheet_name)
        populate_sheet(ws, header, rows)
        print(f"  {sheet_name}: {len(rows)} rows")

    wb.save(OUTPUT_FILE)
    print(f"Excel workbook saved to {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
